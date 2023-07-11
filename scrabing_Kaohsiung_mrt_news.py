import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from datetime import datetime

'''
高雄捷運 新聞稿
'''


def scrabing():

    # 設定高捷新聞稿"主頁面"網址
    url_front = "https://www.krtc.com.tw/"

    list_merge = []

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍 OK
    table = soup.find('div', class_="newsList clearfix")

    li_elems = table.find_all('li')

    for li in li_elems:

        # 網址 後半部
        url_back = li.find("a", class_=None).get('href')

        # 網址前後結合成完整網址
        url = url_front + url_back

        # 標題
        title = li.find("div", class_='word').h3.string

        # 標題底下的部分文章
        article = li.find("div", class_='word').p.string

        # 發布時間
        dt = li.find("div", class_="bot").string
        dt = dt.replace("發布日期：", "")

        # 整合
        list_merge.append([title, dt, article, url])

    print(list_merge)
    return list_merge


def writing_in_csv(list_1):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("高雄捷運新聞稿爬蟲", 0)

    # 儲存檔案
    workbook.save('高雄捷運新聞稿爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('高雄捷運新聞稿爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["高雄捷運新聞稿爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "日期"
    s1["D1"] = "新聞稿預覽"
    s1["E1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in list_1:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]

        # 建立超連結 column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[3]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('高雄捷運新聞稿爬蟲.xlsx')


if __name__ == '__main__':
    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定等待時間上限10秒
    # driver.implicitly_wait(10)

    # 設定高捷新聞稿網址
    url_krtc_mrt = "https://www.krtc.com.tw/Information/news"



    # 指定前往的網頁
    driver.get(url_krtc_mrt)

    list_merge = scrabing()
    # writing_in_csv(list_1=list_merge)
