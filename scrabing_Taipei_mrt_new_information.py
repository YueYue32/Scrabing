import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from datetime import datetime,timedelta


'''
台北市政府捷運工程局 新聞稿
'''

def scrabing():
    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定等待時間上限10秒
    # driver.implicitly_wait(10)

    # 設定網址
    url_taipei_mrt = "https://www.dorts.gov.taipei/News.aspx?n=41977EB83537C82B&sms=72544237BBE4C5F6"

    # 設定"主頁面"網址
    url_original = "https://www.dorts.gov.taipei/"

    # 指定前往的網頁
    driver.get(url_taipei_mrt)

    # 上述整合成二維資料
    merge_2D = []

    # 上述整合成二維資料，後續再轉成字典型態資料
    merge_dict = []
    dict_final = []

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍
    table = soup.find('table', {'id': "table_0"}).tbody

    # 每一列
    rows = table.find_all('tr')

    for row in rows:
        # 標題
        title = row.find_all('td')[1].text

        # 日期
        date = row.find_all('td')[2].string

        # 發布機關
        unit = row.find_all('td')[3].string

        # 後半段網址
        url_back = row.find_all('td')[1].span.a.get('href')

        url = url_original + url_back

        # 年分 日期的前3個字元
        year = int(date[:3])

        # 改成西元年分 (+1911)
        year_vids = year + 1911

        # 組合成西元 日期
        dt_vids = str(year_vids) + date[3:]

        # 轉換成時間格式資料
        dt = datetime.strptime(dt_vids, "%Y-%m-%d")

        # 只要 7天內的 資料
        if dt >= seven_day_ago:
            dt = dt.strftime("%Y-%m-%d")
            merge_2D.append([title, dt, unit, url])
            merge_dict.append([title, dt, unit, url])
        else:
            break

    key = ["標題", "發布時間","發布機關","網址"]

    for merge_part in merge_dict:
        # 鍵值合併
        dict_merge_data = dict(zip(key, merge_part))
        dict_final.append(dict_merge_data)

    print(dict_final)
    # 回傳 二維陣列資料
    return merge_2D

    # 回傳 字典型態資料
    # return dict_final


def writing_in_csv(merge_list):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("台北市捷運爬蟲", 0)

    # 儲存檔案
    workbook.save('台北市捷運爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('台北市捷運爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["台北市捷運爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "發布時間"
    s1["D1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in merge_list:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]

        # 建立超連結 column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[2]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('台北市捷運爬蟲.xlsx')


if __name__ == '__main__':
    # 當天日期
    today = datetime.today()
    # 當天日期的7天前
    seven_day_ago = today - timedelta(days=7)

    list_1 = scrabing()
    # writing_in_csv(merge_list=list_1)

