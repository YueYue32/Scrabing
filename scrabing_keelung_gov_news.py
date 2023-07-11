import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from datetime import datetime
from pynput import mouse
from datetime import datetime
import random

'''
高雄市政府 最新消息
'''


def scrabing():

    list_merge = []

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍 OK
    table = soup.find('div', class_="list")

    # 次範圍 OK
    rows = table.find_all('li')

    for row in rows:
        t1 = row.find("span", class_="list_title").a

        # 標題
        title = t1['title']

        # 網址
        url = t1["href"]

        # 時間
        dt = row.find("time").string

        # 單位
        unit = row.find("span", class_="list_dep").string

        list_merge.append([title, dt, unit, url])


    key = ["標題", "時間", "單位","網址"]

    dict_final = []

    for merge_part in list_merge:
        # 鍵值合併
        dict_merge_data = dict(zip(key, merge_part))
        # 字典型態資料格式
        dict_final.append(dict_merge_data)

    # 回傳 二維陣列資料
    return list_merge

    # 回傳 字典型態資料
    # return dict_final


def writing_in_csv(merge_list):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("基隆市政府最新公告爬蟲", 0)

    # 儲存檔案
    workbook.save('基隆市政府最新公告爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('基隆市政府最新公告爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["基隆市政府最新公告爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "日期"
    s1["D1"] = "發布單位"
    s1["E1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in merge_list:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]

        # 建立超連結 column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[3]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('基隆市政府最新公告爬蟲.xlsx')


if __name__ == '__main__':
    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定等待時間上限10秒
    # driver.implicitly_wait(10)

    # 設定基隆市政府網址
    url_kl_mrt = "https://www.klcg.gov.tw/tw/klcg1/3170.html"

    # 指定前往的網頁
    driver.get(url_kl_mrt)
    list_1 = scrabing()
    # writing_in_csv(merge_list=list_1)
