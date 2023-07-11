import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from datetime import datetime, timedelta

'''
新北市政府 市政新聞
'''

def scrabing():
    list_merge = []

    next_page = driver.find_element(By.CLASS_NAME, "next").is_enabled()

    do_next = True

    while do_next:

        soup = BeautifulSoup(driver.page_source, "html.parser")

        table = soup.find("div", class_="list")

        rows = table.find_all("div", class_="article")

        for row in rows:

            # 標題
            title = row.a.get('title').replace("\u3000", " ").replace("\n", "").strip()

            # 後半段網址
            url_back = row.a.get('href')

            # 合併網址
            url = url_front + url_back

            # top區塊 包含時間、單位
            ti = row.find("div", class_="info_top")

            # 包含時間、單位 的區塊
            upper = ti.find_all("span")

            # 時間
            dt = upper[0].text
            dt_sp = datetime.strptime(dt, "%Y-%m-%d")
            dt_format = datetime.strftime(dt_sp, "%Y/%m/%d")

            # 發布單位
            unit = upper[1].text

            # 抓7天內資料
            if dt_format >= seven_day_ago:
                list_merge.append([title, dt_format, unit, url])

            else:
                do_next = False
                break

        if do_next and next_page:
            driver.find_element(By.CLASS_NAME, "next").click()
            time.sleep(3)
        else:
            break

    key = ["標題", "日期", "發布單位", "網址"]

    dict_final = []

    for merge_part in list_merge:
        # 鍵值合併
        dict_merge_data = dict(zip(key, merge_part))
        # 字典型態資料格式
        dict_final.append(dict_merge_data)

    # 回傳 二維陣列資料
    return list_merge

    # 回傳 字典型態資料
    # return dict_final


def writing_in_csv(list_merge):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("新北市政府市政新聞爬蟲", 0)

    # 儲存檔案
    workbook.save('新北市政府市政新聞爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('新北市政府市政新聞爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["新北市政府市政新聞爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "發布時間"
    s1["D1"] = "發布單位"
    s1["E1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in list_merge:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]

        # 建立超連結 column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[3]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('新北市政府市政新聞爬蟲.xlsx')


if __name__ == '__main__':
    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定等待時間上限10秒
    # driver.implicitly_wait(10)
    main_url = "https://www.ntpc.gov.tw/ch/home.jsp?id=e8ca970cde5c00e1"
    url_front = "https://www.ntpc.gov.tw/ch/"

    # 當天日期
    today = datetime.today()
    # 當天日期的7天前
    seven_day_ago = today - timedelta(days=7)
    seven_day_ago = seven_day_ago.strftime("%Y/%m/%d")

    # 指定前往的網頁
    driver.get(main_url)

    time.sleep(5)

    list_1 = scrabing()
    # writing_in_csv(list_merge=list_1)
