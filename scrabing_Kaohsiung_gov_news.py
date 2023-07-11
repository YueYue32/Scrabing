import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from pynput import mouse
from datetime import datetime, timedelta
import random

'''
高雄市政府全球資訊網 最新消息
'''


def scrabing():

    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定網址
    url_news = "https://www.kcg.gov.tw/News.aspx?n=F29A02A9D36C47F0&sms=19902EF36D6B551D"

    # 設定"主頁面"網址
    url_original = "https://www.kcg.gov.tw/"

    # 指定前往的網頁
    driver.get(url_news)

    # 網頁下方，每頁顯示"X"筆數，網頁的預設值是 "20"，刪掉並改成 "100"
    page_num_in = driver.find_element(By.ID, "input_page")
    page_num_in.click()

    # 按下 ctrl + A 鍵
    page_num_in.send_keys(Keys.CONTROL, "a")
    time.sleep(2)

    # 按下 backspace 鍵
    page_num_in.send_keys(Keys.BACK_SPACE)
    time.sleep(2)

    # 輸入 100
    page_num_in.send_keys("100")

    # 按下 enter 鍵
    page_num_in.send_keys(Keys.ENTER)

    time.sleep(2)

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍
    table = soup.find('table', class_="cell-table table_blue02")

    # tbody
    tbody = table.find("tbody")

    # 用來裝爬取出來的資料
    list1 = []

    rows = tbody.find_all("tr")

    for row in rows:
        part = row.find_all("p", class_=None)

        # 網址的後半
        url_part = row.find('a')['href']

        # 標題
        title = row.find('a')['title']

        # 全角空格字符\u3000，替換成空格
        title = title.replace("\u3000", " ")

        # 和主網址合併
        url = url_original + url_part

        # 發布單位
        unit = part[1].text

        # 發布時間
        dt = part[2].text

        # 發布時間，最前面的數值，改成西元年
        dt_year = dt[0:3]
        dt_year_int = int(dt_year)
        dt_year_int_plus1911 = dt_year_int + 1911
        dt_year_int_plus1911 = str(dt_year_int_plus1911)
        dt_kai = dt_year_int_plus1911 + "-" + dt[4:]
        dt_kai = datetime.strptime(dt_kai, "%Y-%m-%d")
        dt_shinn = dt_kai.strftime("%Y/%m/%d")

        # 當天日期
        today = datetime.today()
        # 當天日期的7天前
        seven_day_ago = today - timedelta(days=7)
        seven_day_ago = seven_day_ago.strftime("%Y/%m/%d")

        # 如果爬出來的發布日期 >= 當天日期的7天前，表示該筆資料的發布日期在"一周內"
        if dt_shinn >= seven_day_ago:
            list1.append([title, unit, dt_shinn, url])

    print(list1)
    return list1


def writing_in_csv(list1):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("高雄市最新消息爬蟲", 0)

    # 儲存檔案
    workbook.save('高雄市最新消息爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('高雄市最新消息爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["高雄市最新消息爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "發布單位"
    s1["D1"] = "發布時間"
    s1["E1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in list1:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]

        # 建立超連結，column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[3]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('高雄市最新消息爬蟲.xlsx')


if __name__ == '__main__':
    list_1 = scrabing()
    # writing_in_csv(list1=list_1)
