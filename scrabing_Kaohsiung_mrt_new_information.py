import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from datetime import datetime

'''
高雄市政府捷運工程局 捷運新訊
'''


def scrabing():
    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定等待時間上限10秒
    # driver.implicitly_wait(10)

    # 設定高捷網址
    url_kaohsiung_mrt = "https://mtbu.kcg.gov.tw/Activities/C002100"

    # 設定高捷"主頁面"網址
    url_original = "https://mtbu.kcg.gov.tw"

    # 指定前往的網頁
    driver.get(url_kaohsiung_mrt)

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍
    table = soup.find('div', class_="main_body")

    # 次範圍
    rows = table.find_all('div', class_="row p-3 p-md-0 align-items-center")

    # 用來裝爬取出來的資料
    list1 = []

    for row in rows:
        # 爬取"標題" OK
        title = row.find("div", class_="col-12 col-md")
        # 改變爬取"title"出來的資料字串內容
        item_new_tltle = title.text.replace("New", "").strip()

        # 爬取"標題的url" OK
        url_title = title.find("a", class_="txt-link")
        url_title = url_title['href']
        url = url_original + url_title

        # 爬取"發布單位" OK
        unit = row.find('div', class_="col-12 col-md-2 text-md-center kf-data-type").text

        # 爬取"時間"
        t = row.find('div', class_="col-12 col-md-2 text-md-center kf-date")
        # 修改時間t(time)的輸出內容(把\n修掉)、修改時間輸出格式
        item_new_time = t.text.replace("\n", "")

        # 爬取"點閱數"
        num_of_click = row.find("div", class_='col-12 col-md-2 text-md-center kf-ctr text-break').text

        # 將爬取資料放入list1
        list1.append([item_new_tltle, unit, item_new_time, num_of_click, url])

    merge_list = list1
    print(merge_list)

    return merge_list


def writing_in_csv(merge_list):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"中央爬蟲測試"
    sheet = workbook.create_sheet("高雄市捷運爬蟲", 0)

    # 儲存檔案
    workbook.save('高雄市捷運爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('高雄市捷運爬蟲.xlsx')

    # 進入工作頁名稱為"爬蟲測試"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["高雄市捷運爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "發布單位"
    s1["D1"] = "發布時間"
    s1["E1"] = "點閱數"
    s1["F1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in merge_list:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]
        s1["F" + str(num)] = data[4]

        # 建立超連結 column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[4]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('高雄市捷運爬蟲.xlsx')


if __name__ == '__main__':
    list_1 = scrabing()
    # writing_in_csv(merge_list=list_1)
