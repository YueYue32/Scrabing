import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from bs4 import BeautifulSoup
import time
import csv
import openpyxl
import pymysql
from pynput import mouse
from datetime import datetime, timedelta
import random
from selenium.webdriver.support.select import Select

'''
台北市政府 市府新聞稿
'''

def scrabing():

    select_category = Select(
                driver.find_element(By.ID,
                                    'SelectNumOnPage'))

    select_category.select_by_value("200")

    time.sleep(2)

    bottun_next = driver.find_element(By.CLASS_NAME, "btn")
    bottun_next.click()

    time.sleep(3)

    list1 = []
    list_title = []
    list_dt = []
    list_unit = []
    list_url = []

    # 使用"BeautifulSoup"，從HTML中分析資料；解析器為"html.parser"，可更換成"html5lib"
    soup = BeautifulSoup(driver.page_source, "html.parser")

    # 最大範圍
    table = soup.find('div', class_="area-table rwd-straight")

    rows_title = table.find_all("td", class_="CCMS_jGridView_td_Class_1")

    rows_time = table.find_all("td", class_="CCMS_jGridView_td_Class_2")

    rows_unit = table.find_all("td", class_="CCMS_jGridView_td_Class_3")

    # 標題區塊
    for row_1 in rows_title:
        part_1 = row_1.span

        # 標題
        title = part_1.a.get('title')

        # 後半部 網址
        url_part = part_1.a.get('href')

        # 合併網址
        url = url_original + url_part
        list_title.append(title)
        list_url.append(url)

    # 日期區塊
    for row_2 in rows_time:

        # 日期
        dt = row_2.span.string

        # 發布時間，最前面的數值，改成西元年
        dt_year = dt[0:3]
        dt_year_int = int(dt_year)
        dt_year_int_plus1911 = dt_year_int + 1911
        dt_year_int_plus1911 = str(dt_year_int_plus1911)
        dt_kai = dt_year_int_plus1911 + dt[3:]
        dt_kai = datetime.strptime(dt_kai, "%Y-%m-%d")
        dt_shinn = dt_kai.strftime("%Y/%m/%d")
        list_dt.append(dt_shinn)

    # 發布單位 區塊
    for row_3 in rows_unit:

        # 單位
        unit = row_3.span.string
        list_unit.append(unit)

    for i in range(len(rows_title)):
        if list_dt[i] >= seven_day_ago:
            list1.append([list_title[i], list_unit[i], list_dt[i], list_url[i]])
    print(list1)
    return list1


def writing_in_csv(list1):
    # 利用 Workbook 建立一個新的工作簿
    workbook = openpyxl.Workbook()

    # 建立一個新工作頁sheet，名稱為"台北市政府新聞稿爬蟲"
    # sheet = workbook.create_sheet("台北市政府新聞稿爬蟲", 0)

    # 儲存檔案
    workbook.save('台北市政府新聞稿爬蟲.xlsx')

    # load指令，進去'openpyxl_test.xlsx'檔案，進入之後令為變數wb
    wb = openpyxl.load_workbook('台北市政府新聞稿爬蟲.xlsx')

    # 進入工作頁名稱為"台北市政府新聞稿爬蟲"的頁面，並且令該頁面為s1，接下來執行動作都在s1上執行
    s1 = wb["台北市政府新聞稿爬蟲"]

    # 寫入標題
    s1["B1"] = "標題"
    s1["C1"] = "發布單位"
    s1["D1"] = "發布時間"
    s1["E1"] = "URL網址"

    # 計數器 count用於編號、num用於列
    count = 0
    num = 2

    for data in list1:
        # 最左邊列-編號
        s1["A" + str(num)] = count + 1
        count += 1

        # 輸入資料，開頭從B列 開始
        s1["B" + str(num)] = data[0]
        s1["C" + str(num)] = data[1]
        s1["D" + str(num)] = data[2]
        s1["E" + str(num)] = data[3]

        # 建立超連結，column是要出現反藍超連結的列數
        s1.cell(row=num, column=2).hyperlink = data[3]        # 連結
        s1.cell(row=num, column=2).value = data[0]            # 匯入超連結的部分(標題)
        s1.cell(row=num, column=2).style = "Hyperlink"        # 模式"Hyperlink"

        num += 1

    # 記得存檔
    wb.save('台北市政府新聞稿爬蟲.xlsx')


if __name__ == '__main__':

    # 建立Chromedriver執行檔位置
    s = Service("chromedriver.exe")

    # 設定瀏覽器為googleChrome
    driver = webdriver.Chrome(service=s)

    # 設定網址
    url_news = "https://www.gov.taipei/News.aspx?n=F0DDAF49B89E9413&sms=72544237BBE4C5F6"

    # 設定"主頁面"網址
    url_original = "https://www.gov.taipei/"

    # 指定前往的網頁
    driver.get(url_news)

    time.sleep(3)

    # 當天日期
    today = datetime.today()
    # 當天日期的7天前
    seven_day_ago = today - timedelta(days=7)
    seven_day_ago = seven_day_ago.strftime("%Y/%m/%d")

    list_1 = scrabing()
    # writing_in_csv(list1=list_1)
